"""
Routes de gestion des candidats
"""
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models import User
from app.services.candidate_service import CandidateService
from app.utils import error_response, candidate_required, admin_required

bp = Blueprint('candidate', __name__)


# ============================================
# ROUTES CANDIDAT (profil personnel)
# ============================================

@bp.route('/profile', methods=['GET'])
@candidate_required()
def get_my_profile():
    """Récupère le profil du candidat connecté"""
    user_id = int(get_jwt_identity())
    
    profile, error = CandidateService.get_profile(user_id)
    
    if error:
        return error_response(error, 404)
    
    return jsonify({
        'success': True,
        'data': profile
    })


@bp.route('/profile', methods=['PUT'])
@candidate_required()
def update_my_profile():
    """
    Met à jour le profil du candidat
    
    Body (tous optionnels):
        - first_name, last_name, birth_date, gender
        - phone, address, city, region
        - parent_name, parent_phone, parent_relation
        - school_name, school_city, class_level
        - average_t1, average_t2, average_t3
    """
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    profile, error = CandidateService.update_profile(user_id, data)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Profil mis à jour',
        'data': profile
    })


@bp.route('/profile/submit', methods=['POST'])
@candidate_required()
def submit_my_profile():
    """Soumet le profil pour validation par un admin"""
    user_id = int(get_jwt_identity())
    
    profile, error = CandidateService.submit_profile(user_id)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Profil soumis pour validation',
        'data': profile
    })


@bp.route('/profile/upload-photo', methods=['POST'])
@candidate_required()
def upload_photo():
    """
    Upload la photo d'identité du candidat (fichier réel)
    
    Body: multipart/form-data
        - file: image (jpg, jpeg, png, webp, avif) max 5MB
    """
    from app.services.file_service import FileService
    
    user_id = int(get_jwt_identity())
    
    if 'file' not in request.files:
        return error_response("Aucun fichier envoyé", 400)
    
    file = request.files['file']
    if file.filename == '':
        return error_response("Nom de fichier vide", 400)
    
    # Sauvegarder l'image
    relative_path, save_error = FileService.save_image(file, subfolder='photos')
    
    if save_error:
        return error_response(save_error, 400)
    
    # Mettre à jour le profil
    photo_url = f'/uploads/{relative_path}'
    update_result, error = CandidateService.upload_photo(user_id, photo_url)
    
    if error:
        FileService.delete_file(relative_path)
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Photo uploadée avec succès',
        'data': {'photo_url': photo_url}
    })


@bp.route('/profile/photo', methods=['POST'])
@candidate_required()
def upload_photo_url():
    """
    Upload photo par URL (rétrocompatibilité)
    
    Body:
        - photo_url: string
    """
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    photo_url = data.get('photo_url')
    if not photo_url:
        return error_response("URL de la photo requise", 400)
    
    result, error = CandidateService.upload_photo(user_id, photo_url)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Photo mise à jour',
        'data': result
    })


@bp.route('/profile/upload-bulletin', methods=['POST'])
@candidate_required()
def upload_bulletin():
    """
    Upload un bulletin scolaire PDF
    
    Body: multipart/form-data
        - file: PDF max 10MB
        - trimestre: int (1, 2 ou 3)
    """
    from app.services.file_service import FileService
    
    user_id = int(get_jwt_identity())
    
    if 'file' not in request.files:
        return error_response("Aucun fichier envoyé", 400)
    
    file = request.files['file']
    if file.filename == '':
        return error_response("Nom de fichier vide", 400)
    
    trimestre = request.form.get('trimestre', type=int)
    if trimestre not in (1, 2, 3):
        return error_response("Le trimestre doit être 1, 2 ou 3", 400)
    
    # Sauvegarder le document
    relative_path, save_error = FileService.save_document(file, subfolder='bulletins')
    
    if save_error:
        return error_response(save_error, 400)
    
    # Mettre à jour le profil
    user = User.query.get(user_id)
    if not user or not user.candidate:
        FileService.delete_file(relative_path)
        return error_response("Profil non trouvé", 404)
    
    candidate = user.candidate
    bulletin_field = f'bulletin_t{trimestre}_url'
    bulletin_url = f'/uploads/{relative_path}'
    
    # Supprimer l'ancien bulletin si présent
    old_url = getattr(candidate, bulletin_field)
    if old_url:
        old_path = old_url.replace('/uploads/', '')
        FileService.delete_file(old_path)
    
    setattr(candidate, bulletin_field, bulletin_url)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'Bulletin trimestre {trimestre} uploadé avec succès',
        'data': {'bulletin_url': bulletin_url, 'trimestre': trimestre}
    })


# ============================================
# ROUTES ADMIN
# ============================================

@bp.route('/admin/list', methods=['GET'])
@admin_required()
def admin_list_candidates():
    """
    Liste tous les candidats avec filtres et pagination
    
    Query params:
        - page: int (default 1)
        - per_page: int (default 20, max 100)
        - status: draft|submitted|validated|rejected
        - region: string
        - gender: M|F
        - class_level: string
        - search: string (nom, email, école)
        - score_min: float
        - score_max: float
        - has_score: yes|no
    """
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    
    filters = {
        'status': request.args.get('status'),
        'region': request.args.get('region'),
        'gender': request.args.get('gender'),
        'class_level': request.args.get('class_level'),
        'search': request.args.get('search'),
        'score_min': request.args.get('score_min', type=float),
        'score_max': request.args.get('score_max', type=float),
        'has_score': request.args.get('has_score')
    }
    
    # Retirer les filtres vides
    filters = {k: v for k, v in filters.items() if v is not None}
    
    result, error = CandidateService.get_all(filters, page, per_page)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'data': result
    })


@bp.route('/admin/<int:candidate_id>', methods=['GET'])
@admin_required()
def admin_get_candidate(candidate_id):
    """Récupère les détails d'un candidat"""
    result, error = CandidateService.get_by_id(candidate_id)
    
    if error:
        return error_response(error, 404)
    
    return jsonify({
        'success': True,
        'data': result
    })


@bp.route('/admin/<int:candidate_id>/validate', methods=['POST'])
@admin_required()
def admin_validate_candidate(candidate_id):
    """
    Valide une candidature
    
    Body (optionnel):
        - comment: string
    """
    admin_id = int(get_jwt_identity())
    data = request.get_json() or {}
    comment = data.get('comment')
    
    result, error = CandidateService.validate(candidate_id, admin_id, comment)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Candidature validée',
        'data': result
    })


@bp.route('/admin/<int:candidate_id>/reject', methods=['POST'])
@admin_required()
def admin_reject_candidate(candidate_id):
    """
    Rejette une candidature
    
    Body:
        - reason: string (obligatoire)
    """
    admin_id = int(get_jwt_identity())
    data = request.get_json() or {}
    reason = data.get('reason', '')
    
    result, error = CandidateService.reject(candidate_id, admin_id, reason)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Candidature rejetée',
        'data': result
    })


@bp.route('/admin/bulk-validate', methods=['POST'])
@admin_required()
def admin_bulk_validate():
    """
    Valide plusieurs candidatures
    
    Body:
        - candidate_ids: array of int
    """
    admin_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    candidate_ids = data.get('candidate_ids', [])
    if not candidate_ids:
        return error_response("Liste de candidats requise", 400)
    
    result, error = CandidateService.bulk_validate(candidate_ids, admin_id)
    
    return jsonify({
        'success': True,
        'message': f"{result['total_validated']} candidature(s) validée(s)",
        'data': result
    })


@bp.route('/admin/stats', methods=['GET'])
@admin_required()
def admin_candidate_stats():
    """Statistiques des candidats"""
    result, error = CandidateService.get_stats()
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'data': result
    })


@bp.route('/admin/export', methods=['GET'])
@admin_required()
def admin_export_candidates():
    """
    Exporte la liste des candidats
    
    Query params: 
        - format: json|csv|xlsx (défaut: json)
        - Mêmes filtres que /admin/list
    """
    import io
    import csv
    from flask import Response
    
    export_format = request.args.get('format', 'json').lower()
    
    filters = {
        'status': request.args.get('status'),
        'region': request.args.get('region'),
        'gender': request.args.get('gender'),
        'search': request.args.get('search')
    }
    filters = {k: v for k, v in filters.items() if v is not None}
    
    # Récupérer tous sans pagination
    result, error = CandidateService.get_all(filters, page=1, per_page=10000)
    
    if error:
        return error_response(error, 400)
    
    candidates = result['candidates']
    
    if export_format == 'json':
        return jsonify({
            'success': True,
            'data': candidates,
            'total': result['total']
        })
    
    # Colonnes pour CSV/Excel
    headers = [
        'ID', 'Prénom', 'Nom', 'Email', 'Genre', 'Date de naissance',
        'Téléphone', 'Ville', 'Région', 'École', 'Classe',
        'Moyenne T1', 'Moyenne T2', 'Moyenne T3', 'Moyenne Générale',
        'Moyenne Maths', 'Moyenne Sciences',
        'Statut', 'Score QCM', 'Date inscription'
    ]
    
    def extract_row(c):
        return [
            c.get('id', ''),
            c.get('first_name', ''),
            c.get('last_name', ''),
            c.get('email', ''),
            c.get('gender', ''),
            c.get('birth_date', ''),
            c.get('phone', ''),
            c.get('city', ''),
            c.get('region', ''),
            c.get('school_name', ''),
            c.get('class_level', ''),
            c.get('academic', {}).get('average_t1', '') if isinstance(c.get('academic'), dict) else c.get('average_t1', ''),
            c.get('academic', {}).get('average_t2', '') if isinstance(c.get('academic'), dict) else c.get('average_t2', ''),
            c.get('academic', {}).get('average_t3', '') if isinstance(c.get('academic'), dict) else c.get('average_t3', ''),
            c.get('academic', {}).get('average_score', '') if isinstance(c.get('academic'), dict) else c.get('average_score', ''),
            c.get('academic', {}).get('math_average', '') if isinstance(c.get('academic'), dict) else c.get('math_average', ''),
            c.get('academic', {}).get('science_average', '') if isinstance(c.get('academic'), dict) else c.get('science_average', ''),
            c.get('status', ''),
            c.get('qcm_score', ''),
            c.get('created_at', '')
        ]
    
    if export_format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for c in candidates:
            writer.writerow(extract_row(c))
        
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=candidats_olympiades.csv'}
        )
    
    elif export_format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidats"
        
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="206080", end_color="206080", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_idx, c in enumerate(candidates, 2):
            for col_idx, val in enumerate(extract_row(c), 1):
                ws.cell(row=row_idx, column=col_idx, value=val if val != '' else None)
        
        # Largeurs auto
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=candidats_olympiades.xlsx'}
        )
