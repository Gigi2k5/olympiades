"""
Routes du QCM
"""
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models import Question
from app.services.qcm_service import QCMService, QCMAdminService
from app.utils import error_response, candidate_required, admin_required

bp = Blueprint('qcm', __name__)


# ============================================
# ROUTES CANDIDAT
# ============================================

@bp.route('/status', methods=['GET'])
@candidate_required()
def get_qcm_status():
    """Vérifie le statut du QCM pour le candidat"""
    user_id = int(get_jwt_identity())
    result, error = QCMService.get_attempt_status(user_id)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({'success': True, 'data': result})


@bp.route('/settings', methods=['GET'])
@jwt_required()
def get_qcm_settings():
    """Récupère les paramètres publics du QCM"""
    result, error = QCMService.get_settings()
    
    # Ne retourner que les infos publiques
    public_settings = {
        'duration_minutes': result['duration_minutes'],
        'total_questions': result['total_questions'],
        'is_open': result['is_open']
    }
    
    return jsonify({'success': True, 'data': public_settings})


@bp.route('/start', methods=['POST'])
@candidate_required()
def start_qcm():
    """Démarre le QCM"""
    user_id = int(get_jwt_identity())
    result, error = QCMService.start_qcm(user_id)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'QCM démarré',
        'data': result
    })


@bp.route('/answer', methods=['POST'])
@candidate_required()
def save_answer():
    """
    Sauvegarde une réponse
    
    Body:
        - attempt_id: int
        - question_index: int
        - answer_index: int (0-3)
    """
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    attempt_id = data.get('attempt_id')
    question_index = data.get('question_index')
    answer_index = data.get('answer_index')
    
    if attempt_id is None or question_index is None or answer_index is None:
        return error_response("attempt_id, question_index et answer_index requis", 400)
    
    result, error = QCMService.save_answer(user_id, attempt_id, question_index, answer_index)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({'success': True, 'data': result})


@bp.route('/submit', methods=['POST'])
@candidate_required()
def submit_qcm():
    """
    Soumet le QCM
    
    Body:
        - attempt_id: int
    """
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    attempt_id = data.get('attempt_id')
    if not attempt_id:
        return error_response("attempt_id requis", 400)
    
    result, error = QCMService.submit_qcm(user_id, attempt_id)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'QCM soumis avec succès',
        'data': result
    })


@bp.route('/result', methods=['GET'])
@candidate_required()
def get_result():
    """Récupère le résultat du QCM"""
    user_id = int(get_jwt_identity())
    result, error = QCMService.get_result(user_id)
    
    if error:
        return error_response(error, 404)
    
    return jsonify({'success': True, 'data': result})


# ============================================
# ROUTES ADMIN - QUESTIONS
# ============================================

@bp.route('/admin/questions', methods=['GET'])
@admin_required()
def admin_list_questions():
    """Liste toutes les questions"""
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    
    filters = {
        'category': request.args.get('category'),
        'difficulty': request.args.get('difficulty'),
        'is_active': request.args.get('is_active', type=lambda x: x.lower() == 'true') if request.args.get('is_active') else None,
        'search': request.args.get('search')
    }
    filters = {k: v for k, v in filters.items() if v is not None}
    
    result, error = QCMAdminService.get_all_questions(filters, page, per_page)
    
    return jsonify({'success': True, 'data': result})


@bp.route('/admin/questions', methods=['POST'])
@admin_required()
def admin_create_question():
    """
    Crée une question
    
    Body:
        - text: string
        - option_a, option_b, option_c, option_d: string
        - correct_answer: int (0-3)
        - category: string
        - difficulty: easy|medium|hard
    """
    admin_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    result, error = QCMAdminService.create_question(data, admin_id)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Question créée',
        'data': result
    }), 201


@bp.route('/admin/questions/<int:question_id>', methods=['PUT'])
@admin_required()
def admin_update_question(question_id):
    """Met à jour une question"""
    admin_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    result, error = QCMAdminService.update_question(question_id, data, admin_id)
    
    if error:
        return error_response(error, 404)
    
    return jsonify({
        'success': True,
        'message': 'Question mise à jour',
        'data': result
    })


@bp.route('/admin/questions/<int:question_id>', methods=['DELETE'])
@admin_required()
def admin_delete_question(question_id):
    """Supprime une question"""
    result, error = QCMAdminService.delete_question(question_id)
    
    if error:
        return error_response(error, 404)
    
    return jsonify({
        'success': True,
        'message': 'Question supprimée'
    })


# ============================================
# ROUTES ADMIN - PARAMÈTRES
# ============================================

@bp.route('/admin/settings', methods=['GET'])
@admin_required()
def admin_get_settings():
    """Récupère tous les paramètres"""
    result, error = QCMService.get_settings()
    return jsonify({'success': True, 'data': result})


@bp.route('/admin/settings', methods=['PUT'])
@admin_required()
def admin_update_settings():
    """Met à jour les paramètres"""
    admin_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    result, error = QCMAdminService.update_settings(data, admin_id)
    
    if error:
        return error_response(error, 400)
    
    return jsonify({
        'success': True,
        'message': 'Paramètres mis à jour',
        'data': result
    })


@bp.route('/admin/stats', methods=['GET'])
@admin_required()
def admin_qcm_stats():
    """Statistiques du QCM"""
    result, error = QCMAdminService.get_qcm_stats()
    return jsonify({'success': True, 'data': result})


# ============================================
# ANTI-TRICHE
# ============================================

@bp.route('/report-event', methods=['POST'])
@candidate_required()
def report_cheat_event():
    """
    Enregistre un événement anti-triche
    
    Body:
        - attempt_id: int
        - event_type: string (tab_switch, fullscreen_exit, copy_attempt, right_click)
        - timestamp: string ISO (optionnel)
    """
    from app.models import QCMAttempt
    
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    
    attempt_id = data.get('attempt_id')
    event_type = data.get('event_type')
    
    if not attempt_id or not event_type:
        return error_response("attempt_id et event_type requis", 400)
    
    valid_events = ['tab_switch', 'fullscreen_exit', 'copy_attempt', 'right_click']
    if event_type not in valid_events:
        return error_response(f"event_type invalide. Valeurs acceptées: {', '.join(valid_events)}", 400)
    
    # Vérifier que la tentative appartient au candidat
    from app.models import User
    user = User.query.get(user_id)
    if not user or not user.candidate:
        return error_response("Candidat non trouvé", 404)
    
    attempt = QCMAttempt.query.filter_by(
        id=attempt_id,
        candidate_id=user.candidate.id
    ).first()
    
    if not attempt:
        return error_response("Tentative non trouvée", 404)
    
    if attempt.status != 'in_progress':
        return error_response("La tentative n'est plus en cours", 400)
    
    # Enregistrer l'événement
    attempt.add_cheat_event(event_type)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'data': {
            'tab_switches': attempt.tab_switches or 0,
            'fullscreen_exits': attempt.fullscreen_exits or 0,
            'is_flagged': attempt.is_flagged or False
        }
    })


# ============================================
# IMPORT / EXPORT QUESTIONS
# ============================================

@bp.route('/admin/questions/export', methods=['GET'])
@admin_required()
def admin_export_questions():
    """
    Exporte les questions
    
    Query params:
        - format: json|xlsx (défaut: json)
    """
    import io
    import json as json_lib
    from flask import Response
    from app.models import Question
    
    export_format = request.args.get('format', 'json').lower()
    
    questions = Question.query.order_by(Question.id).all()
    questions_data = [q.to_dict(include_answer=True) for q in questions]
    
    if export_format == 'json':
        # Export JSON propre (pas d'enveloppe API, juste les données)
        output = json_lib.dumps(questions_data, indent=2, ensure_ascii=False)
        return Response(
            output,
            mimetype='application/json',
            headers={'Content-Disposition': 'attachment; filename=questions_olympiades.json'}
        )
    
    elif export_format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"
        
        headers = ['ID', 'Texte', 'Option A', 'Option B', 'Option C', 'Option D',
                    'Réponse correcte', 'Catégorie', 'Difficulté', 'Actif',
                    'Fois affichée', 'Fois correcte', 'Taux réussite']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="206080", end_color="206080", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, q in enumerate(questions, 2):
            ws.cell(row=row_idx, column=1, value=q.id)
            ws.cell(row=row_idx, column=2, value=q.text)
            ws.cell(row=row_idx, column=3, value=q.option_a)
            ws.cell(row=row_idx, column=4, value=q.option_b)
            ws.cell(row=row_idx, column=5, value=q.option_c)
            ws.cell(row=row_idx, column=6, value=q.option_d)
            ws.cell(row=row_idx, column=7, value=q.correct_answer)
            ws.cell(row=row_idx, column=8, value=q.category)
            ws.cell(row=row_idx, column=9, value=q.difficulty)
            ws.cell(row=row_idx, column=10, value='Oui' if q.is_active else 'Non')
            ws.cell(row=row_idx, column=11, value=q.times_shown)
            ws.cell(row=row_idx, column=12, value=q.times_correct)
            ws.cell(row=row_idx, column=13, value=q.success_rate)
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=questions_olympiades.xlsx'}
        )
    
    return error_response("Format non supporté. Utilisez json ou xlsx.", 400)


@bp.route('/admin/questions/import', methods=['POST'])
@admin_required()
def admin_import_questions():
    """
    Importe des questions en masse
    
    Body: multipart/form-data
        - file: fichier JSON ou Excel (.json, .xlsx, .xls)
    
    Format JSON attendu: array de {text, option_a, option_b, option_c, option_d, correct_answer, category, difficulty}
    Format Excel attendu: colonnes text, option_a, option_b, option_c, option_d, correct_answer, category, difficulty
    """
    import json as json_lib
    from app.models import Question, AuditLog
    
    admin_id = int(get_jwt_identity())
    
    if 'file' not in request.files:
        return error_response("Aucun fichier envoyé", 400)
    
    file = request.files['file']
    filename = file.filename.lower()
    
    questions_data = []
    
    try:
        if filename.endswith('.json'):
            content = file.read().decode('utf-8')
            questions_data = json_lib.loads(content)
            if not isinstance(questions_data, list):
                return error_response("Le JSON doit contenir un tableau de questions", 400)
        
        elif filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            import io
            
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            
            # Lire les en-têtes de la première ligne
            headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
            
            required = ['text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer', 'category', 'difficulty']
            
            # Mapper les colonnes
            col_map = {}
            for req in required:
                for idx, h in enumerate(headers):
                    if h == req or h.replace(' ', '_') == req:
                        col_map[req] = idx
                        break
            
            missing_cols = [r for r in required if r not in col_map]
            if missing_cols:
                return error_response(f"Colonnes manquantes dans Excel: {', '.join(missing_cols)}", 400)
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                values = [cell.value for cell in row]
                if not values[col_map['text']]:
                    continue
                questions_data.append({
                    req: values[col_map[req]] for req in required
                })
        
        else:
            return error_response("Format non supporté. Utilisez .json, .xlsx ou .xls", 400)
        
    except Exception as e:
        return error_response(f"Erreur de lecture du fichier: {str(e)}", 400)
    
    # Valider et importer
    imported = 0
    errors = []
    valid_categories = ['Logique', 'Algorithmique', 'Mathématiques', 'IA', 'Culture Générale']
    valid_difficulties = ['easy', 'medium', 'hard']
    
    for idx, q in enumerate(questions_data, 1):
        row_errors = []
        
        if not q.get('text'):
            row_errors.append("Texte manquant")
        if not q.get('option_a'):
            row_errors.append("Option A manquante")
        if not q.get('option_b'):
            row_errors.append("Option B manquante")
        if not q.get('option_c'):
            row_errors.append("Option C manquante")
        if not q.get('option_d'):
            row_errors.append("Option D manquante")
        
        try:
            correct = int(q.get('correct_answer', -1))
            if correct not in (0, 1, 2, 3):
                row_errors.append("correct_answer doit être 0, 1, 2 ou 3")
        except (ValueError, TypeError):
            row_errors.append("correct_answer invalide")
            correct = -1
        
        category = str(q.get('category', '')).strip()
        difficulty = str(q.get('difficulty', 'medium')).strip().lower()
        
        if difficulty not in valid_difficulties:
            row_errors.append(f"Difficulté invalide: {difficulty}")
        
        if row_errors:
            errors.append({'row': idx, 'errors': row_errors})
            continue
        
        question = Question(
            text=str(q['text']).strip(),
            option_a=str(q['option_a']).strip(),
            option_b=str(q['option_b']).strip(),
            option_c=str(q['option_c']).strip(),
            option_d=str(q['option_d']).strip(),
            correct_answer=correct,
            category=category if category else 'Logique',
            difficulty=difficulty,
            created_by=admin_id,
            is_active=True
        )
        db.session.add(question)
        imported += 1
    
    if imported > 0:
        AuditLog.log(
            user_id=admin_id,
            action='import_questions',
            entity_type='question',
            details=f"{imported} questions importées"
        )
        db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f"{imported} question(s) importée(s)",
        'data': {
            'imported': imported,
            'errors': errors,
            'total_in_file': len(questions_data)
        }
    })
